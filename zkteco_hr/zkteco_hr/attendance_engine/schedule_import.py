"""
Spreadsheet-based bulk schedule import.

Supports xlsx (and csv) files with the format:
  Col 0: Employee ID card (e.g. DI-0159)
  Col 1: Email
  Col 2: AM From  (e.g. 7:30am)
  Col 3: AM To    (e.g. 12:00pm)
  Col 4: PM From  (e.g. 1:00pm, or "off")
  Col 5: PM To    (e.g. 5:00pm)
  Col 6: Day off  (e.g. "Saturday & Sunday", "Sat(Afternoon) and Sunday")

Exposed API: parse_schedule_upload(file_b64, filename)
"""

from __future__ import annotations

import base64
import io
import re

import frappe

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
_WEEKDAY_IDX: dict[str, int] = {d: i for i, d in enumerate(WEEKDAYS)}

_DAY_ALIASES: dict[str, str] = {
    "mon": "Monday",
    "monday": "Monday",
    "tue": "Tuesday",
    "tues": "Tuesday",
    "tuesday": "Tuesday",
    "wed": "Wednesday",
    "wednesday": "Wednesday",
    "thu": "Thursday",
    "thur": "Thursday",
    "thurs": "Thursday",
    "thursday": "Thursday",
    "fri": "Friday",
    "friday": "Friday",
    "sat": "Saturday",
    "saturday": "Saturday",
    "sun": "Sunday",
    "sunday": "Sunday",
}

_TIME_RE = re.compile(r"^(\d{1,2}):(\d{2})\s*(am|pm)?$", re.IGNORECASE)
_ID_PATTERN = re.compile(r"^[A-Za-z]{1,4}-\d{2,}", re.IGNORECASE)

# ---------------------------------------------------------------------------
# Day-range helpers
# ---------------------------------------------------------------------------


def _normalize_weekday(s: str) -> str | None:
    return _DAY_ALIASES.get(s.lower().strip())


def _expand_day_range(start: str, end: str) -> list[str]:
    s = _normalize_weekday(start)
    e = _normalize_weekday(end)
    if not s or not e:
        return []
    si, ei = _WEEKDAY_IDX[s], _WEEKDAY_IDX[e]
    return WEEKDAYS[si : ei + 1] if si <= ei else []


def _parse_day_off_text(text) -> dict:
    """
    Parse day-off description into structured off-day sets.

    Returns {"full_off": list[str], "afternoon_off": list[str]}

    Handles patterns like:
    - "Saturday & Sunday"                          → full_off: [Sat, Sun]
    - "Sat(Afternoon) and Sunday"                  → full_off: [Sun], afternoon_off: [Sat]
    - "Mon-Fri(Afternoon)Saturday & Sunday"        → full_off: [Sat, Sun], afternoon_off: [Mon..Fri]
    """
    if not text or str(text).strip().lower() in ("nan", "none", ""):
        return {"full_off": [], "afternoon_off": []}

    text = str(text).strip()
    full_off: set[str] = set()
    afternoon_off: set[str] = set()

    # Extract "Day(Afternoon)" and "DayRange(Afternoon)" patterns first
    pm_re = re.compile(r"([A-Za-z]+-[A-Za-z]+|[A-Za-z]+)\s*\(Afternoon\)", re.IGNORECASE)
    remaining = text
    for m in pm_re.finditer(text):
        remaining = remaining.replace(m.group(0), "")
        part = m.group(1)
        if "-" in part:
            s, e = part.split("-", 1)
            afternoon_off.update(_expand_day_range(s.strip(), e.strip()))
        else:
            d = _normalize_weekday(part)
            if d:
                afternoon_off.add(d)

    # Remaining text → full off days
    for token in re.split(r"[&,+]|\band\b", remaining, flags=re.IGNORECASE):
        token = token.strip()
        if not token:
            continue
        if "-" in token:
            parts = token.split("-", 1)
            if _normalize_weekday(parts[0].strip()) and _normalize_weekday(parts[1].strip()):
                full_off.update(_expand_day_range(parts[0].strip(), parts[1].strip()))
        else:
            d = _normalize_weekday(token)
            if d:
                full_off.add(d)

    return {
        "full_off": sorted(full_off, key=lambda d: _WEEKDAY_IDX[d]),
        "afternoon_off": sorted(afternoon_off, key=lambda d: _WEEKDAY_IDX[d]),
    }


# ---------------------------------------------------------------------------
# Time parsing
# ---------------------------------------------------------------------------


def _is_off(value) -> bool:
    if value is None:
        return True
    return str(value).strip().lower() in ("nan", "none", "", "off")


def _parse_time(value) -> str | None:
    if _is_off(value):
        return None
    text = str(value).strip()
    m = _TIME_RE.match(text)
    if not m:
        return None
    h, mi, period = int(m.group(1)), int(m.group(2)), (m.group(3) or "").lower()
    if period == "pm" and h != 12:
        h += 12
    elif period == "am" and h == 12:
        h = 0
    return f"{h:02d}:{mi:02d}"


def _cell_str(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() in ("nan", "none") else s


# ---------------------------------------------------------------------------
# File parsing
# ---------------------------------------------------------------------------


def _parse_xlsx_bytes(file_bytes: bytes) -> list[list]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def _parse_csv_bytes(file_bytes: bytes) -> list[list]:
    import csv

    text = file_bytes.decode("utf-8-sig", errors="replace")
    return list(csv.reader(io.StringIO(text)))


def _find_data_start(rows: list[list]) -> int:
    """Return the index of the first row that looks like an employee data row."""
    for i, row in enumerate(rows):
        if row and row[0] is not None and _ID_PATTERN.match(str(row[0]).strip()):
            return i
    return len(rows)


# ---------------------------------------------------------------------------
# Employee lookup
# ---------------------------------------------------------------------------


def _lookup_employee(id_card: str, email: str = "") -> tuple[str | None, str | None]:
    """Return (employee_name, display_name) or (None, None)."""
    for field in ("employee_number", "attendance_device_id"):
        result = frappe.db.get_value(
            "Employee",
            {field: id_card, "status": "Active"},
            ["name", "employee_name"],
            as_dict=True,
        )
        if result:
            return result["name"], result["employee_name"]

    if email:
        for email_field in ("company_email", "personal_email"):
            result = frappe.db.get_value(
                "Employee",
                {email_field: email, "status": "Active"},
                ["name", "employee_name"],
                as_dict=True,
            )
            if result:
                return result["name"], result["employee_name"]

    return None, None


# ---------------------------------------------------------------------------
# WeekPattern builder
# ---------------------------------------------------------------------------


def _build_week_pattern(
    am_from: str | None,
    am_to: str | None,
    pm_from: str | None,
    pm_to: str | None,
    day_off: dict,
) -> dict:
    """
    Build a WeekPattern dict suitable for apply_weekly_schedule.

    Logic:
    - full_off days → works: False
    - afternoon_off days OR pm=off → works AM times only (no lunch block)
    - Otherwise → works full day: start=am_from, end=pm_to, lunch=am_to→pm_from
    """
    full_off = set(day_off.get("full_off") or [])
    afternoon_off = set(day_off.get("afternoon_off") or [])
    has_pm = bool(pm_from and pm_to)

    days = []
    for weekday in WEEKDAYS:
        if weekday in full_off:
            days.append({"weekday": weekday, "works": False})
            continue

        if weekday in afternoon_off or not has_pm:
            days.append(
                {
                    "weekday": weekday,
                    "works": True,
                    "start_time": am_from,
                    "end_time": am_to,
                    "lunch_start": None,
                    "lunch_end": None,
                    "grace_minutes": 10,
                }
            )
        else:
            days.append(
                {
                    "weekday": weekday,
                    "works": True,
                    "start_time": am_from,
                    "end_time": pm_to,
                    "lunch_start": am_to,
                    "lunch_end": pm_from,
                    "grace_minutes": 10,
                }
            )

    return {"frequency": "Every Week", "days": days}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


@frappe.whitelist()
def parse_schedule_upload(file_b64: str, filename: str = "upload.xlsx") -> dict:
    """
    Parse a base64-encoded xlsx/csv schedule file and return a structured preview.

    Each row in the result contains:
      id_card, email, employee (Frappe name), employee_name, matched (bool),
      am_from, am_to, pm_from, pm_to, day_off, week_pattern, warnings
    """
    try:
        file_bytes = base64.b64decode(file_b64)
    except Exception as exc:
        frappe.throw(f"Invalid file data: {exc}")

    ext = (filename or "").lower().rsplit(".", 1)[-1]
    try:
        raw_rows = _parse_csv_bytes(file_bytes) if ext == "csv" else _parse_xlsx_bytes(file_bytes)
    except Exception as exc:
        frappe.throw(f"Failed to read spreadsheet: {exc}")

    data_start = _find_data_start(raw_rows)
    if data_start >= len(raw_rows):
        frappe.throw(
            "No employee rows detected. "
            "Expected rows where the first column is an employee ID (e.g. DI-0159)."
        )

    result_rows = []
    for raw in raw_rows[data_start:]:
        # Pad to at least 7 columns
        while len(raw) < 7:
            raw.append(None)

        id_card = _cell_str(raw[0])
        if not id_card:
            continue

        email = _cell_str(raw[1])
        am_from = _parse_time(raw[2])
        am_to = _parse_time(raw[3])
        pm_from = _parse_time(raw[4]) if not _is_off(raw[4]) else None
        pm_to = _parse_time(raw[5]) if not _is_off(raw[5]) else None
        day_off = _parse_day_off_text(raw[6])

        warnings: list[str] = []
        if not am_from or not am_to:
            warnings.append("AM shift times missing or unparseable")
        if not _is_off(raw[4]) and (pm_from is None or pm_to is None):
            warnings.append(f"PM time unparseable: {_cell_str(raw[4])!r} – {_cell_str(raw[5])!r}")

        employee, employee_name = _lookup_employee(id_card, email)
        if not employee:
            warnings.append(f"No active employee found for ID card {id_card!r}")

        week_pattern = None
        if am_from and am_to:
            week_pattern = _build_week_pattern(am_from, am_to, pm_from, pm_to, day_off)

        result_rows.append(
            {
                "id_card": id_card,
                "email": email,
                "employee": employee,
                "employee_name": employee_name,
                "matched": bool(employee),
                "am_from": am_from,
                "am_to": am_to,
                "pm_from": pm_from,
                "pm_to": pm_to,
                "day_off": day_off,
                "week_pattern": week_pattern,
                "warnings": warnings,
            }
        )

    return {"rows": result_rows}
